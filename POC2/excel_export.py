"""
Build a multi-sheet XLSX workbook from a POC2 ExtractionResult.

Why this lives in its own module:
  - The UI (app.py) is already long; keeping the workbook layout here keeps
    the Streamlit code readable.
  - openpyxl is a heavier import than the rest of POC2 — isolating it means
    the CLI / extractor path doesn't pay for it.

The workbook has four sheets:
  1. "Extractions"    — flat tabular view of the canonical rows the UI shows.
                         Uses verified rows when verification ran, else raw
                         extractions. Matches what the JSON downloads expose.
  2. "Coverage"       — 37 rows: Metric | Found (Yes/No).
  3. "Summary"        — key/value pairs from result.totals + the consolidated-
                         filter stats. Two columns: Key | Value.
  4. "Per-Metric Log" — one row per metric call: name, status, kept rows,
                         attempts, elapsed seconds, token usage. Lets analysts
                         spot retries / null returns at a glance.
"""
from __future__ import annotations

import io
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="2E3440")
_WRAP = Alignment(wrap_text=True, vertical="top")


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet, max_width: int = 60) -> None:
    """Cheap autosizing — measure max printed length per column, clamp."""
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        longest = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            # Newlines blow up the visual width; measure the widest line.
            longest = max(longest, max(len(line) for line in str(v).splitlines() or [""]))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(
            max(12, longest + 2), max_width
        )


def _stringify(v: Any) -> Any:
    """Excel rejects raw lists/dicts. Stringify those; pass primitives through."""
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

_EXTRACTION_COLS: list[str] = [
    "metric_target",
    "current_year_value",
    "declared_unit",
    "entity_context",
    "source_type",
    "page_number",
    "table_or_section",
    "company_definition_quote",
    "verbatim_source_text",
    "forensic_reasoning_log",
    "verified",
    "verification_note",
]


def _write_extractions(ws: Worksheet, rows: Iterable[dict]) -> None:
    _write_header(ws, _EXTRACTION_COLS)
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, key in enumerate(_EXTRACTION_COLS, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=_stringify(row.get(key)))
            if key in ("verbatim_source_text", "forensic_reasoning_log",
                       "verification_note", "table_or_section", "company_definition_quote"):
                cell.alignment = _WRAP
    _autosize(ws)


def _write_coverage(ws: Worksheet, coverage: dict[str, bool]) -> None:
    _write_header(ws, ["Metric", "Found"])
    for r_idx, (name, found) in enumerate(coverage.items(), start=2):
        ws.cell(row=r_idx, column=1, value=name)
        ws.cell(row=r_idx, column=2, value="Yes" if found else "No")
    _autosize(ws)


def _write_summary(ws: Worksheet, result) -> None:
    _write_header(ws, ["Key", "Value"])
    items: list[tuple[str, Any]] = [
        ("Company", result.company_display),
        ("FY year", result.fy_year),
        ("Model", result.model),
    ]
    for k, v in result.totals.items():
        items.append((k, _stringify(v)))
    for r_idx, (k, v) in enumerate(items, start=2):
        ws.cell(row=r_idx, column=1, value=k)
        ws.cell(row=r_idx, column=2, value=v)
    _autosize(ws)


_LOG_COLS: list[str] = [
    "metric",
    "status",
    "raw_row_count",
    "kept_row_count",
    "attempts",
    "elapsed_s",
    "input_tokens",
    "output_tokens",
    "cached_tokens",
    "total_tokens",
]


def _write_per_metric_log(ws: Worksheet, log: list[dict]) -> None:
    _write_header(ws, _LOG_COLS)
    for r_idx, entry in enumerate(log, start=2):
        usage = entry.get("usage") or {}
        values = [
            entry.get("metric"),
            entry.get("status"),
            entry.get("raw_row_count"),
            entry.get("kept_row_count"),
            entry.get("attempts"),
            entry.get("elapsed_s"),
            usage.get("input_tokens"),
            usage.get("output_tokens"),
            usage.get("cached_tokens"),
            usage.get("total_tokens"),
        ]
        for c_idx, v in enumerate(values, start=1):
            ws.cell(row=r_idx, column=c_idx, value=_stringify(v))
    _autosize(ws)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def build_excel_workbook(result) -> bytes:
    """Return the .xlsx bytes for the given ExtractionResult."""
    wb = Workbook()

    # The default sheet ships first — repurpose it as "Extractions".
    ws = wb.active
    ws.title = "Extractions"
    _write_extractions(ws, result.verified or result.extractions)

    _write_coverage(wb.create_sheet("Coverage"), result.coverage)
    _write_summary(wb.create_sheet("Summary"), result)
    _write_per_metric_log(wb.create_sheet("Per-Metric Log"), result.per_metric_log)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
