#!/usr/bin/env python3
"""
Megasheet Merger: Scans the PDF directory for all generated Excel files,
extracts the findings, and merges them into two stacked databases:
- Consolidated Database (Matrix & Flat formats)
- Standalone Database (Matrix & Flat formats)
"""
import os
import sys
import glob
import argparse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from pathlib import Path

# All 37 target metrics in order
ALL_METRICS = [
    "Operating Income", "Adjusted Revenue", "Adjusted Earnings", "EBIT", "EBITDA",
    "Adjusted EBIT", "Core Earnings", "Normalized Earnings", "Recurring Earnings",
    "Adjusted EPS", "Normalized EPS", "GAAP One-Time Adjusted", "GAAP Adjusted",
    "Free Cash Flow", "Funds from Operations (FFO)", "Distributable Cash Flow",
    "Net Debt", "Cash Earnings", "Cash Loss", "Constant-Currency Revenues",
    "Constant-Currency Revenue Growth", "Constant-Currency Operating Expenses",
    "PPOP", "Pre-sales", "Bookings", "Collection Value", "Economic Value Added (EVA)",
    "Credit Cost ex one-off", "Adjusted ROE", "Adjusted ROA", "Adjusted Book Value",
    "Operating Margin", "Adjusted EBITDA Margin", "Adjusted EBIT Margin",
    "Adjusted PBT Margin", "Adjusted Net Profit Margin", "Base Business Margin"
]

def parse_args():
    parser = argparse.ArgumentParser(description="Merge all extraction sheets into consolidated and standalone databases.")
    parser.add_argument("--dir", required=True, help="Base directory containing company subfolders (e.g. ./pdfs)")
    parser.add_argument("--suffix", default="_POC3", help="Excel suffix to look for (e.g. _POC3, _POC3_batch3)")
    parser.add_argument("--out", default="merged_database.xlsx", help="Filename of the output merged workbook")
    return parser.parse_args()

def extract_from_workbook(path):
    wb = openpyxl.load_workbook(path)
    company = Path(path).parent.name
    
    # Try to derive year from filename e.g. 13_POC3.xlsx -> FY13
    stem = Path(path).stem
    year_part = stem.split("_")[0]
    fy_year = f"FY{year_part[-2:]}" if year_part.isdigit() else "Unknown"

    cons_rows = []
    std_rows = []

    # Process Consolidated
    if "Consolidated Disclosures" in wb.sheetnames:
        ws = wb["Consolidated Disclosures"]
        for row in list(ws.iter_rows(min_row=2, values_only=True)):
            if not row or not row[0]:
                continue
            metric = row[0]
            val = row[1]
            proof = row[3]
            src_type = row[4]
            page = row[5]
            verb = row[9]
            # Convert values to float/int if possible
            val_clean = clean_val(val)
            cons_rows.append({
                "company": company, "year": fy_year, "metric": metric, "value": val_clean,
                "proof": proof, "source_type": src_type, "page": page, "verbatim": verb
            })

    # Process Standalone
    if "Standalone Disclosures" in wb.sheetnames:
        ws = wb["Standalone Disclosures"]
        for row in list(ws.iter_rows(min_row=2, values_only=True)):
            if not row or not row[0]:
                continue
            metric = row[0]
            val = row[1]
            proof = row[3]
            src_type = row[4]
            page = row[5]
            verb = row[9]
            val_clean = clean_val(val)
            std_rows.append({
                "company": company, "year": fy_year, "metric": metric, "value": val_clean,
                "proof": proof, "source_type": src_type, "page": page, "verbatim": verb
            })

    return cons_rows, std_rows

def clean_val(val):
    if val is None:
        return None
    s = str(val).strip()
    if any(kw in s.lower() for kw in ["0 candidates", "rejected all", "none", "-", "not_incurred"]):
        return s
    
    # Remove percent signs, commas, and try to convert to float/int
    s_clean = s.replace("%", "").replace(",", "").replace("₹", "").replace("Rs", "").strip()
    try:
        if "." in s_clean:
            return float(s_clean)
        return int(s_clean)
    except ValueError:
        return s

def build_matrix_sheet(wb, sheet_name, data_rows, header_fill, border_side):
    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True

    # 1. Set up Headers
    headers = ["Company Name", "Fiscal Year"] + ALL_METRICS
    ws.append(headers)

    # Styling headers
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=border_side)

    # 2. Pivot data by (company, year)
    pivoted = {}
    for r in data_rows:
        key = (r["company"], r["year"])
        pivoted.setdefault(key, {})[r["metric"]] = r["value"]

    # 3. Write rows sorted by Company, Year
    row_idx = 2
    for (comp, yr), metrics in sorted(pivoted.items()):
        ws.cell(row=row_idx, column=1, value=comp).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=2, value=yr).alignment = Alignment(horizontal="center")
        
        for col_idx, metric_name in enumerate(ALL_METRICS, 3):
            val = metrics.get(metric_name, "-")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="center")
        row_idx += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def build_flat_sheet(wb, sheet_name, data_rows, header_fill, border_side):
    ws = wb.create_sheet(title=sheet_name)
    ws.views.sheetView[0].showGridLines = True

    headers = ["Company Name", "Fiscal Year", "Metric Name", "Extracted Value", "Page Number", "Source Type", "Verbatim Proof", "Scope Conviction Proof"]
    ws.append(headers)

    # Style Headers
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=border_side)

    # Write Data
    row_idx = 2
    for r in sorted(data_rows, key=lambda x: (x["company"], x["year"], x["metric"])):
        ws.cell(row=row_idx, column=1, value=r["company"]).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=2, value=r["year"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=3, value=r["metric"]).alignment = Alignment(horizontal="left")
        
        val_cell = ws.cell(row=row_idx, column=4, value=r["value"])
        if isinstance(r["value"], (int, float)):
            val_cell.number_format = "#,##0.00"
            val_cell.alignment = Alignment(horizontal="right")
        else:
            val_cell.alignment = Alignment(horizontal="center")

        ws.cell(row=row_idx, column=5, value=r["page"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=6, value=r["source_type"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=7, value=r["verbatim"]).alignment = Alignment(horizontal="left")
        ws.cell(row=row_idx, column=8, value=r["proof"]).alignment = Alignment(horizontal="left")
        row_idx += 1

    # Set Widths
    for col in ws.columns:
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        if col[0].column in (7, 8):
            ws.column_dimensions[col_letter].width = 40
        else:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

def main():
    args = parse_args()
    base_path = Path(args.dir).resolve()
    
    if not base_path.exists() or not base_path.is_dir():
        print(f"Error: Base directory not found: {base_path}")
        sys.exit(1)

    print(f"=== INITIALIZING MEGASHEET MERGING ===")
    print(f"Directory:     {base_path}")
    print(f"Suffix Filter: {args.suffix}")
    print(f"Output File:   {args.out}")
    print(f"======================================\n")

    # Find all matching files
    search_pattern = os.path.join(base_path, f"**/*{args.suffix}.xlsx")
    matched_files = sorted(glob.glob(search_pattern, recursive=True))

    # Exclude any previously generated merged_database.xlsx
    matched_files = [f for f in matched_files if Path(f).name != Path(args.out).name]

    if not matched_files:
        print(f"No Excel files matching '*{args.suffix}.xlsx' found.")
        sys.exit(0)

    print(f"Found {len(matched_files)} output workbooks to merge.")

    all_cons_rows = []
    all_std_rows = []

    for path in matched_files:
        try:
            c_rows, s_rows = extract_from_workbook(path)
            all_cons_rows.extend(c_rows)
            all_std_rows.extend(s_rows)
            print(f"  * Merged {Path(path).parent.name} ({Path(path).name})")
        except Exception as e:
            print(f"  * ❌ Failed to merge {path}: {e}")

    # Create the output workbook
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Styles
    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    dark_green_fill = PatternFill(start_color="2A5C2A", end_color="2A5C2A", fill_type="solid")
    thin_border = Side(border_style="thin", color="CCCCCC")

    print("\nWriting Consolidated sheets...")
    build_matrix_sheet(wb, "Consolidated Matrix", all_cons_rows, navy_fill, thin_border)
    build_flat_sheet(wb, "Consolidated Master Flat", all_cons_rows, navy_fill, thin_border)

    print("Writing Standalone sheets...")
    build_matrix_sheet(wb, "Standalone Matrix", all_std_rows, dark_green_fill, thin_border)
    build_flat_sheet(wb, "Standalone Master Flat", all_std_rows, dark_green_fill, thin_border)

    out_path = base_path / args.out
    wb.save(out_path)
    print(f"\n=======================================================")
    print(f"🎉 SUCCESS: Database Merged & Stacked!")
    print(f"Merged File Path: {out_path}")
    print(f"Total Consolidated Entries: {len(all_cons_rows)}")
    print(f"Total Standalone Entries:   {len(all_std_rows)}")
    print(f"=======================================================")

if __name__ == "__main__":
    main()
