"""
Excel exporter for POC3: Two-Stage Exhaustive Candidate Extraction & LLM Finalization Layer.

Generates a multi-sheet openpyxl workbook:
  - Sheet 1: Finalized Disclosures (winning values and locations)
  - Sheet 2: Candidate Audit Log (every candidate harvested across the PDF in Layer 1 and its details)
  - Sheet 3: Coverage & Stats (summary of metrics found/missing and token usage)
"""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from POC3.metrics import METRIC_METADATA


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="Calibri", size=11)
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
TOP_ALIGN = Alignment(vertical="top")


def _style_headers(ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int = 1) -> None:
    for cell in ws[row_idx]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row_idx].height = 26


def _auto_fit_columns(ws: openpyxl.worksheet.worksheet.Worksheet, max_len_cap: int = 50) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = min(len(val_str), max_len_cap)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def _populate_disclosures_sheet(ws, metrics_list: list[dict], title: str) -> None:
    """Helper to populate a disclosures worksheet (Consolidated or Standalone)."""
    ws.title = title
    ws.views.sheetView[0].showGridLines = True

    headers = [
        "Metric Target", "Final Value", "Entity Context", "Scope Conviction Proof", "Source Type",
        "Page #", "Printed Page", "Table / Section",
        "Audit Log Summary", "Verbatim Source Text"
    ]
    ws.append(headers)
    _style_headers(ws)

    for item in metrics_list:
        val = item.get("final_value")
        win = item.get("winning_candidate") or {}
        audit_log = item.get("rejection_audit_log", [])
        audit_str = "\n".join(audit_log) if isinstance(audit_log, list) else str(audit_log)

        if val is not None:
            disp_val = val
        elif "0 candidate" in audit_str or "Zero candidate" in audit_str or "0 Consolidated" in audit_str or "0 Standalone" in audit_str:
            disp_val = "0 CANDIDATES"
        else:
            disp_val = "REJECTED ALL"

        row = [
            item.get("metric_target", ""),
            disp_val,
            win.get("entity_context", "-") if win else "-",
            win.get("scope_conviction_proof", "-") if win else "-",
            win.get("source_type", "-") if win else "-",
            win.get("page_number", "") if win else "",
            win.get("printed_page_number", "") if win else "",
            win.get("table_or_section", "") if win else "",
            audit_str,
            win.get("verbatim_source_text", "") if win else "",
        ]
        ws.append(row)
        row_idx = ws.max_row
        for col_idx in range(1, len(row) + 1):
            c = ws.cell(row=row_idx, column=col_idx)
            c.font = BOLD_FONT if col_idx == 2 and val is not None else DATA_FONT
            c.border = THIN_BORDER
            c.alignment = WRAP_ALIGN if col_idx in (9, 10) else TOP_ALIGN

    _auto_fit_columns(ws, max_len_cap=60)


def export_to_excel(result: Any, output_path: Path | str) -> None:
    """Export an ExtractionResultPOC3 to a 4-sheet Excel workbook (`Consolidated`, `Standalone`, `Audit Log`, `Stats`)."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Consolidated Disclosures ────────────────────────────────────
    ws1 = wb.active
    cons_metrics = getattr(result, "finalized_consolidated_metrics", getattr(result, "finalized_metrics", []))
    _populate_disclosures_sheet(ws1, cons_metrics, "Consolidated Disclosures")

    # ── Sheet 2: Standalone Disclosures ──────────────────────────────────────
    ws2 = wb.create_sheet(title="Standalone Disclosures")
    std_metrics = getattr(result, "finalized_standalone_metrics", [])
    _populate_disclosures_sheet(ws2, std_metrics, "Standalone Disclosures")

    # ── Sheet 3: Candidate Audit Log ─────────────────────────────────────────
    ws3 = wb.create_sheet(title="Candidate Audit Log")
    ws3.views.sheetView[0].showGridLines = True

    headers3 = [
        "Metric Target", "Candidate Value", "Entity Context", "Scope Conviction Proof", "Source Type",
        "Page #", "Printed Page", "Table / Section",
        "Line Above Proof", "Line Below Proof",
        "Verbatim Text", "Forensic Reasoning Log"
    ]
    ws3.append(headers3)
    _style_headers(ws3)

    harvested = getattr(result, "harvested_candidates", {})
    for metric_name, cand_list in harvested.items():
        if not cand_list:
            row = [metric_name, "NO CANDIDATES FOUND", "", "", "", "", "", "", "", "", "", ""]
            ws3.append(row)
            continue
        for cand in cand_list:
            row = [
                cand.get("metric_target", metric_name),
                cand.get("current_year_value", ""),
                cand.get("entity_context", ""),
                cand.get("scope_conviction_proof", ""),
                cand.get("source_type", ""),
                cand.get("page_number", ""),
                cand.get("printed_page_number", ""),
                cand.get("table_or_section", ""),
                cand.get("page_verbatim_proof_above", ""),
                cand.get("page_verbatim_proof_below", ""),
                cand.get("verbatim_source_text", ""),
                cand.get("forensic_reasoning_log", ""),
            ]
            ws3.append(row)
            row_idx = ws3.max_row
            for col_idx in range(1, len(row) + 1):
                c = ws3.cell(row=row_idx, column=col_idx)
                c.font = DATA_FONT
                c.border = THIN_BORDER
                c.alignment = WRAP_ALIGN if col_idx in (9, 10, 11, 12) else TOP_ALIGN

    _auto_fit_columns(ws3, max_len_cap=50)

    # ── Sheet 4: Coverage & Stats ────────────────────────────────────────────
    ws4 = wb.create_sheet(title="Coverage & Stats")
    ws4.views.sheetView[0].showGridLines = True

    headers4 = [
        "Metric Name", "Consolidated Disclosed?", "Consolidated Value",
        "Standalone Disclosed?", "Standalone Value", "Total Candidates Harvested"
    ]
    ws4.append(headers4)
    _style_headers(ws4)

    cons_cov = getattr(result, "consolidated_coverage", getattr(result, "coverage", {}))
    std_cov = getattr(result, "standalone_coverage", {})
    cons_map = {m.get("metric_target"): m.get("final_value") for m in cons_metrics}
    std_map = {m.get("metric_target"): m.get("final_value") for m in std_metrics}

    for m in METRIC_METADATA:
        mname = m["name"]
        disc_cons = cons_cov.get(mname, False)
        disc_std = std_cov.get(mname, False)
        cand_count = len(harvested.get(mname, []))

        # Consolidated Value display
        fval_cons = cons_map.get(mname)
        if fval_cons is not None:
            disp_cons = fval_cons
        elif cand_count == 0:
            disp_cons = "0 CANDIDATES"
        else:
            disp_cons = "REJECTED ALL"

        # Standalone Value display
        fval_std = std_map.get(mname)
        if fval_std is not None:
            disp_std = fval_std
        elif cand_count == 0:
            disp_std = "0 CANDIDATES"
        else:
            disp_std = "REJECTED ALL"

        row = [mname, "YES" if disc_cons else "NO", disp_cons, "YES" if disc_std else "NO", disp_std, cand_count]
        ws4.append(row)
        row_idx = ws4.max_row
        for col_idx in range(1, len(row) + 1):
            c = ws4.cell(row=row_idx, column=col_idx)
            c.font = BOLD_FONT if (col_idx == 3 and disc_cons) or (col_idx == 5 and disc_std) else DATA_FONT
            c.border = THIN_BORDER
            c.alignment = TOP_ALIGN

    _auto_fit_columns(ws4, max_len_cap=40)

    wb.save(output_path)

